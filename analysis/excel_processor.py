# analysis/excel_processor.py

"""
Excel Processing and Automation Utilities for Framework0.

This module provides comprehensive Excel processing capabilities including:
- Data cleaning and transformation operations
- Analysis and reporting features  
- Visualization and dashboard creation
- Multi-sheet workbook management with navigation

All functions follow Framework0 standards with full typing, logging integration,
and modular design for reusability across automation workflows.

Features:
- Data cleaning: Remove duplicates, standardize formats, clean raw exports
- Analysis: Generate pivot tables, filter data, create summaries
- Visualization: Auto-generate charts, conditional formatting, dashboards
- Navigation: Table of contents and sheet navigation for large workbooks
"""

import os
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.hyperlink import Hyperlink
from typing import Dict, List, Optional, Union, Any, Tuple
from pathlib import Path
import logging
from datetime import datetime
import re

# Import Framework0 logger
from src.core.logger import get_logger

# Initialize module logger with debug support
logger = get_logger(__name__, debug=os.getenv("DEBUG") == "1")


class ExcelProcessorV1:
    """
    Excel processing and automation engine for Framework0.
    
    Provides comprehensive Excel operations including data cleaning,
    analysis, visualization, and multi-sheet workbook management.
    Follows Framework0 patterns with modular design and full logging.
    """
    
    def __init__(self, filepath: str, *, debug: bool = False) -> None:
        """
        Initialize Excel processor with target workbook.
        
        Args:
            filepath (str): Path to Excel file to process
            debug (bool): Enable debug logging for operations
        """
        # Initialize instance attributes with proper typing
        self.filepath: str = filepath  # Target Excel file path
        self.debug: bool = debug  # Debug mode flag
        self.workbook: Optional[Workbook] = None  # Loaded workbook instance
        self.original_sheets: List[str] = []  # Original sheet names for reference
        
        # Setup logger with debug configuration
        self.logger = get_logger(f"{__name__}.{self.__class__.__name__}", debug=debug)
        self.logger.info(f"Excel processor initialized for file: {filepath}")
    
    def load_workbook(self) -> Workbook:
        """
        Load Excel workbook from filepath, creating new if doesn't exist.
        
        Returns:
            Workbook: Loaded or created openpyxl Workbook instance
            
        Raises:
            FileNotFoundError: If filepath directory doesn't exist
            PermissionError: If file access is denied
        """
        try:
            # Check if file exists and load, otherwise create new workbook
            if Path(self.filepath).exists():
                self.logger.debug(f"Loading existing workbook: {self.filepath}")
                self.workbook = load_workbook(self.filepath)
                # Store original sheet names for reference
                self.original_sheets = self.workbook.sheetnames.copy()
                self.logger.info(f"Loaded workbook with {len(self.original_sheets)} sheets")
            else:
                self.logger.debug(f"Creating new workbook: {self.filepath}")
                self.workbook = Workbook()
                # Remove default sheet if we're creating new content
                if 'Sheet' in self.workbook.sheetnames:
                    self.workbook.remove(self.workbook['Sheet'])
                self.logger.info("Created new empty workbook")
            
            return self.workbook
            
        except Exception as e:
            self.logger.error(f"Failed to load/create workbook: {e}")
            raise
    
    def save_workbook(self, output_path: Optional[str] = None) -> str:
        """
        Save workbook to file, using original path if no output specified.
        
        Args:
            output_path (Optional[str]): Custom save path, uses self.filepath if None
            
        Returns:
            str: Path where workbook was saved
            
        Raises:
            ValueError: If workbook not loaded
            PermissionError: If file access denied
        """
        if not self.workbook:
            raise ValueError("No workbook loaded - call load_workbook() first")
        
        # Use provided path or default to original filepath
        save_path = output_path or self.filepath
        
        try:
            # Ensure directory exists
            Path(save_path).parent.mkdir(parents=True, exist_ok=True)
            
            self.logger.debug(f"Saving workbook to: {save_path}")
            self.workbook.save(save_path)
            self.logger.info(f"✅ Workbook saved successfully: {save_path}")
            
            return save_path
            
        except Exception as e:
            self.logger.error(f"❌ Failed to save workbook: {e}")
            raise

    def remove_duplicates_from_sheet(self, sheet_name: str, *, 
                                   columns: Optional[List[str]] = None,
                                   keep_first: bool = True) -> int:
        """
        Remove duplicate rows from specified worksheet.
        
        Args:
            sheet_name (str): Name of worksheet to process
            columns (Optional[List[str]]): Columns to check for duplicates, all if None
            keep_first (bool): Keep first occurrence of duplicates vs last
            
        Returns:
            int: Number of duplicate rows removed
            
        Raises:
            KeyError: If sheet doesn't exist
            ValueError: If workbook not loaded
        """
        if not self.workbook:
            raise ValueError("No workbook loaded - call load_workbook() first")
        
        if sheet_name not in self.workbook.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found in workbook")
        
        self.logger.debug(f"Removing duplicates from sheet: {sheet_name}")
        
        # Get worksheet and convert to DataFrame for processing
        worksheet = self.workbook[sheet_name]
        
        # Extract data from worksheet to DataFrame
        data = []
        headers = []
        
        # Get headers from first row
        for cell in worksheet[1]:
            headers.append(cell.value if cell.value is not None else "")
        
        # Get all data rows
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            data.append(list(row))
        
        if not data:
            self.logger.info("No data found in sheet - no duplicates to remove")
            return 0
        
        # Create DataFrame for duplicate processing
        df = pd.DataFrame(data, columns=headers)
        original_count = len(df)
        
        # Remove duplicates based on specified columns
        subset_cols = columns if columns else None
        df_deduplicated = df.drop_duplicates(subset=subset_cols, keep='first' if keep_first else 'last')
        
        removed_count = original_count - len(df_deduplicated)
        
        if removed_count > 0:
            # Clear existing data and write deduplicated data back
            worksheet.delete_rows(2, worksheet.max_row)
            
            # Write deduplicated data back to worksheet
            for row_data in dataframe_to_rows(df_deduplicated, index=False, header=False):
                worksheet.append(row_data)
            
            self.logger.info(f"✅ Removed {removed_count} duplicate rows from sheet '{sheet_name}'")
        else:
            self.logger.info(f"No duplicates found in sheet '{sheet_name}'")
        
        return removed_count

    def standardize_date_formats(self, sheet_name: str, 
                                date_columns: List[str],
                                target_format: str = "MM/DD/YYYY") -> int:
        """
        Standardize date formats in specified columns of a worksheet.
        
        Args:
            sheet_name (str): Name of worksheet to process
            date_columns (List[str]): Column names containing dates to standardize
            target_format (str): Target Excel date format pattern
            
        Returns:
            int: Number of date cells standardized
            
        Raises:
            KeyError: If sheet or columns don't exist
            ValueError: If workbook not loaded
        """
        if not self.workbook:
            raise ValueError("No workbook loaded - call load_workbook() first")
        
        if sheet_name not in self.workbook.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found in workbook")
        
        worksheet = self.workbook[sheet_name]
        self.logger.debug(f"Standardizing date formats in sheet: {sheet_name}")
        
        # Convert to DataFrame for easier date processing
        data = []
        headers = []
        
        # Get headers
        for cell in worksheet[1]:
            headers.append(cell.value if cell.value is not None else "")
        
        # Find column indices for date columns
        date_col_indices = []
        for col_name in date_columns:
            if col_name in headers:
                date_col_indices.append(headers.index(col_name))
            else:
                self.logger.warning(f"Date column '{col_name}' not found in sheet")
        
        if not date_col_indices:
            self.logger.warning("No valid date columns found")
            return 0
        
        standardized_count = 0
        
        # Process each row starting from row 2 (after headers)
        for row_num in range(2, worksheet.max_row + 1):
            for col_idx in date_col_indices:
                cell = worksheet.cell(row=row_num, column=col_idx + 1)
                
                if cell.value is not None:
                    try:
                        # Try to parse the date value
                        if isinstance(cell.value, str):
                            # Parse string dates with pandas
                            parsed_date = pd.to_datetime(cell.value, infer_datetime_format=True)
                            cell.value = parsed_date
                        
                        # Apply Excel number format to the cell
                        cell.number_format = target_format
                        standardized_count += 1
                        
                    except Exception as e:
                        self.logger.debug(f"Could not parse date in row {row_num}, col {col_idx}: {e}")
                        continue
        
        self.logger.info(f"✅ Standardized {standardized_count} date cells in sheet '{sheet_name}'")
        return standardized_count

    def clean_text_casing(self, sheet_name: str, 
                         text_columns: List[str],
                         case_type: str = "title") -> int:
        """
        Standardize text casing in specified columns.
        
        Args:
            sheet_name (str): Name of worksheet to process
            text_columns (List[str]): Column names to standardize casing
            case_type (str): Type of casing - 'title', 'upper', 'lower', 'proper'
            
        Returns:
            int: Number of text cells processed
            
        Raises:
            KeyError: If sheet doesn't exist
            ValueError: If workbook not loaded or invalid case_type
        """
        if not self.workbook:
            raise ValueError("No workbook loaded - call load_workbook() first")
        
        if sheet_name not in self.workbook.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found in workbook")
        
        valid_cases = ['title', 'upper', 'lower', 'proper']
        if case_type not in valid_cases:
            raise ValueError(f"Invalid case_type '{case_type}'. Must be one of: {valid_cases}")
        
        worksheet = self.workbook[sheet_name]
        self.logger.debug(f"Cleaning text casing in sheet: {sheet_name}, case_type: {case_type}")
        
        # Get headers to find column indices
        headers = []
        for cell in worksheet[1]:
            headers.append(cell.value if cell.value is not None else "")
        
        # Find column indices for text columns
        text_col_indices = []
        for col_name in text_columns:
            if col_name in headers:
                text_col_indices.append(headers.index(col_name))
            else:
                self.logger.warning(f"Text column '{col_name}' not found in sheet")
        
        if not text_col_indices:
            self.logger.warning("No valid text columns found")
            return 0
        
        processed_count = 0
        
        # Process each row starting from row 2 (after headers)
        for row_num in range(2, worksheet.max_row + 1):
            for col_idx in text_col_indices:
                cell = worksheet.cell(row=row_num, column=col_idx + 1)
                
                if cell.value is not None and isinstance(cell.value, str):
                    original_value = cell.value.strip()  # Remove leading/trailing spaces
                    
                    if original_value:  # Only process non-empty strings
                        if case_type == "title":
                            cell.value = original_value.title()
                        elif case_type == "upper":
                            cell.value = original_value.upper()
                        elif case_type == "lower":
                            cell.value = original_value.lower()
                        elif case_type == "proper":
                            # Proper case - capitalize first letter of each word, but be smarter about it
                            cell.value = " ".join(word.capitalize() for word in original_value.split())
                        
                        processed_count += 1
        
        self.logger.info(f"✅ Processed {processed_count} text cells in sheet '{sheet_name}' with {case_type} casing")
        return processed_count

    def normalize_column_names(self, sheet_name: str) -> Dict[str, str]:
        """
        Normalize column header names for consistency.
        
        Args:
            sheet_name (str): Name of worksheet to process
            
        Returns:
            Dict[str, str]: Mapping of original names to normalized names
            
        Raises:
            KeyError: If sheet doesn't exist
            ValueError: If workbook not loaded
        """
        if not self.workbook:
            raise ValueError("No workbook loaded - call load_workbook() first")
        
        if sheet_name not in self.workbook.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found in workbook")
        
        worksheet = self.workbook[sheet_name]
        self.logger.debug(f"Normalizing column names in sheet: {sheet_name}")
        
        name_mapping = {}
        
        # Process header row (row 1)
        for col_num in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col_num)
            
            if cell.value is not None:
                original_name = str(cell.value).strip()
                
                if original_name:
                    # Normalize the column name
                    # Remove extra spaces, convert to title case, remove special chars
                    normalized_name = re.sub(r'[^\w\s]', '', original_name)  # Remove special chars
                    normalized_name = re.sub(r'\s+', ' ', normalized_name)  # Collapse multiple spaces
                    normalized_name = normalized_name.strip().title()  # Title case and trim
                    
                    # Replace common abbreviations and standardize
                    replacements = {
                        'Id': 'ID',
                        'Qty': 'Quantity', 
                        'Amt': 'Amount',
                        'Desc': 'Description',
                        'Num': 'Number',
                        'Addr': 'Address',
                        'Tel': 'Telephone',
                        'Email': 'Email',
                        'Fax': 'Fax'
                    }
                    
                    for old, new in replacements.items():
                        normalized_name = normalized_name.replace(old, new)
                    
                    # Only update if different
                    if normalized_name != original_name:
                        name_mapping[original_name] = normalized_name
                        cell.value = normalized_name
                        self.logger.debug(f"Normalized column: '{original_name}' -> '{normalized_name}'")
        
        if name_mapping:
            self.logger.info(f"✅ Normalized {len(name_mapping)} column names in sheet '{sheet_name}'")
        else:
            self.logger.info(f"No column names required normalization in sheet '{sheet_name}'")
        
        return name_mapping

    def split_sheet_by_column(self, sheet_name: str, 
                             split_column: str, 
                             prefix: str = "Sheet") -> Dict[str, str]:
        """
        Split a worksheet into multiple sheets based on unique values in a column.
        
        Args:
            sheet_name (str): Source sheet name to split
            split_column (str): Column name to split on
            prefix (str): Prefix for new sheet names
            
        Returns:
            Dict[str, str]: Mapping of split values to new sheet names
            
        Raises:
            KeyError: If sheet or column doesn't exist
            ValueError: If workbook not loaded
        """
        if not self.workbook:
            raise ValueError("No workbook loaded - call load_workbook() first")
        
        if sheet_name not in self.workbook.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found in workbook")
        
        worksheet = self.workbook[sheet_name]
        self.logger.debug(f"Splitting sheet '{sheet_name}' by column '{split_column}'")
        
        # Convert sheet to DataFrame for easier processing
        data = []
        headers = []
        
        # Get headers
        for cell in worksheet[1]:
            headers.append(cell.value if cell.value is not None else "")
        
        if split_column not in headers:
            raise KeyError(f"Split column '{split_column}' not found in sheet")
        
        # Get all data
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            data.append(list(row))
        
        if not data:
            self.logger.warning("No data found to split")
            return {}
        
        # Create DataFrame
        df = pd.DataFrame(data, columns=headers)
        
        # Get unique values from split column
        unique_values = df[split_column].dropna().unique()
        sheet_mapping = {}
        
        for value in unique_values:
            # Create safe sheet name
            safe_value = str(value)[:25]  # Excel sheet names limited to 31 chars
            safe_value = re.sub(r'[^\w\s-]', '', safe_value)  # Remove invalid chars
            new_sheet_name = f"{prefix}_{safe_value}"
            
            # Ensure unique sheet name
            counter = 1
            original_name = new_sheet_name
            while new_sheet_name in self.workbook.sheetnames:
                new_sheet_name = f"{original_name}_{counter}"
                counter += 1
            
            # Filter data for this value
            filtered_df = df[df[split_column] == value]
            
            # Create new worksheet
            new_worksheet = self.workbook.create_sheet(title=new_sheet_name)
            
            # Write headers
            for col_idx, header in enumerate(headers, 1):
                new_worksheet.cell(row=1, column=col_idx, value=header)
            
            # Write filtered data
            for row_idx, row_data in enumerate(dataframe_to_rows(filtered_df, index=False, header=False), 2):
                for col_idx, cell_value in enumerate(row_data, 1):
                    new_worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
            
            sheet_mapping[str(value)] = new_sheet_name
            self.logger.debug(f"Created sheet '{new_sheet_name}' with {len(filtered_df)} rows")
        
        self.logger.info(f"✅ Split sheet into {len(sheet_mapping)} new sheets")
        return sheet_mapping

    def merge_sheets(self, sheet_names: List[str], 
                    target_sheet_name: str, 
                    remove_source: bool = False) -> int:
        """
        Merge multiple sheets into a single sheet.
        
        Args:
            sheet_names (List[str]): Names of sheets to merge
            target_sheet_name (str): Name for merged sheet
            remove_source (bool): Remove source sheets after merge
            
        Returns:
            int: Total number of rows in merged sheet
            
        Raises:
            KeyError: If any source sheet doesn't exist
            ValueError: If workbook not loaded or sheets have different structures
        """
        if not self.workbook:
            raise ValueError("No workbook loaded - call load_workbook() first")
        
        # Validate all sheets exist
        missing_sheets = [name for name in sheet_names if name not in self.workbook.sheetnames]
        if missing_sheets:
            raise KeyError(f"Sheets not found: {missing_sheets}")
        
        self.logger.debug(f"Merging {len(sheet_names)} sheets into '{target_sheet_name}'")
        
        merged_data = []
        headers = None
        
        for sheet_name in sheet_names:
            worksheet = self.workbook[sheet_name]
            
            # Get headers from first sheet
            if headers is None:
                headers = []
                for cell in worksheet[1]:
                    headers.append(cell.value if cell.value is not None else "")
            else:
                # Verify headers match
                current_headers = []
                for cell in worksheet[1]:
                    current_headers.append(cell.value if cell.value is not None else "")
                
                if current_headers != headers:
                    self.logger.warning(f"Sheet '{sheet_name}' has different headers - skipping")
                    continue
            
            # Get data rows from sheet
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):  # Skip empty rows
                    merged_data.append(list(row))
        
        if not merged_data:
            self.logger.warning("No data found to merge")
            return 0
        
        # Create or get target worksheet
        if target_sheet_name in self.workbook.sheetnames:
            target_worksheet = self.workbook[target_sheet_name]
            # Clear existing content
            target_worksheet.delete_rows(1, target_worksheet.max_row)
        else:
            target_worksheet = self.workbook.create_sheet(title=target_sheet_name)
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            target_worksheet.cell(row=1, column=col_idx, value=header)
        
        # Write merged data
        for row_idx, row_data in enumerate(merged_data, 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                target_worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Remove source sheets if requested
        if remove_source:
            for sheet_name in sheet_names:
                if sheet_name != target_sheet_name:  # Don't remove target if it was a source
                    self.workbook.remove(self.workbook[sheet_name])
                    self.logger.debug(f"Removed source sheet: {sheet_name}")
        
        total_rows = len(merged_data) + 1  # +1 for header
        self.logger.info(f"✅ Merged {len(sheet_names)} sheets into '{target_sheet_name}' with {total_rows} total rows")
        
        return total_rows

    def create_table_of_contents(self, toc_sheet_name: str = "Table of Contents") -> str:
        """
        Create a table of contents sheet with navigation links to all worksheets.
        
        Args:
            toc_sheet_name (str): Name for the table of contents sheet
            
        Returns:
            str: Name of created TOC sheet
            
        Raises:
            ValueError: If workbook not loaded
        """
        if not self.workbook:
            raise ValueError("No workbook loaded - call load_workbook() first")
        
        self.logger.debug(f"Creating table of contents: {toc_sheet_name}")
        
        # Create or clear TOC sheet
        if toc_sheet_name in self.workbook.sheetnames:
            toc_sheet = self.workbook[toc_sheet_name]
            toc_sheet.delete_rows(1, toc_sheet.max_row)
        else:
            toc_sheet = self.workbook.create_sheet(title=toc_sheet_name, index=0)  # First sheet
        
        # Set up title styling
        title_font = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_alignment = Alignment(horizontal='center', vertical='center')
        
        header_font = Font(name='Calibri', size=12, bold=True)
        header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        
        # Create title
        toc_sheet.cell(row=1, column=1, value="TABLE OF CONTENTS")
        toc_sheet.cell(row=1, column=1).font = title_font
        toc_sheet.cell(row=1, column=1).fill = title_fill
        toc_sheet.cell(row=1, column=1).alignment = title_alignment
        
        # Merge title cells
        toc_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        
        # Create headers
        headers = ["Sheet Name", "Description", "Row Count", "Last Modified"]
        for col_idx, header in enumerate(headers, 1):
            cell = toc_sheet.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Add sheet information
        row_idx = 4
        for sheet_name in self.workbook.sheetnames:
            if sheet_name != toc_sheet_name:  # Don't include TOC itself
                worksheet = self.workbook[sheet_name]
                
                # Create hyperlink to sheet
                link_cell = toc_sheet.cell(row=row_idx, column=1, value=sheet_name)
                link_cell.hyperlink = f"#{sheet_name}!A1"
                link_cell.font = Font(name='Calibri', size=11, color="0563C1", underline="single")
                
                # Add description (empty for now - could be enhanced)
                toc_sheet.cell(row=row_idx, column=2, value="Data Sheet")
                
                # Add row count
                row_count = worksheet.max_row if worksheet.max_row > 1 else 0
                toc_sheet.cell(row=row_idx, column=3, value=row_count)
                
                # Add current date as last modified
                toc_sheet.cell(row=row_idx, column=4, value=datetime.now().strftime("%Y-%m-%d"))
                
                row_idx += 1
        
        # Auto-adjust column widths
        for column in toc_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            toc_sheet.column_dimensions[column_letter].width = adjusted_width
        
        self.logger.info(f"✅ Created table of contents with {row_idx - 4} sheet links")
        return toc_sheet_name

    def add_navigation_buttons(self, toc_sheet_name: str = "Table of Contents") -> int:
        """
        Add navigation buttons to all sheets linking back to table of contents.
        
        Args:
            toc_sheet_name (str): Name of the table of contents sheet
            
        Returns:
            int: Number of navigation buttons added
            
        Raises:
            ValueError: If workbook not loaded
            KeyError: If TOC sheet doesn't exist
        """
        if not self.workbook:
            raise ValueError("No workbook loaded - call load_workbook() first")
        
        if toc_sheet_name not in self.workbook.sheetnames:
            raise KeyError(f"TOC sheet '{toc_sheet_name}' not found - create TOC first")
        
        self.logger.debug(f"Adding navigation buttons linking to '{toc_sheet_name}'")
        
        button_count = 0
        
        for sheet_name in self.workbook.sheetnames:
            if sheet_name != toc_sheet_name:  # Don't add button to TOC itself
                worksheet = self.workbook[sheet_name]
                
                # Find a good position for the button (top-right corner)
                button_col = min(worksheet.max_column + 2, 26)  # Don't go past column Z
                button_row = 1
                
                # Create button cell
                button_cell = worksheet.cell(row=button_row, column=button_col, value="TOC")
                
                # Style the button
                button_font = Font(name='Calibri', size=10, bold=True, color="FFFFFF")
                button_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                button_alignment = Alignment(horizontal='center', vertical='center')
                button_border = Border(
                    left=Side(border_style="thin"),
                    right=Side(border_style="thin"), 
                    top=Side(border_style="thin"),
                    bottom=Side(border_style="thin")
                )
                
                button_cell.font = button_font
                button_cell.fill = button_fill
                button_cell.alignment = button_alignment
                button_cell.border = button_border
                
                # Add hyperlink to TOC
                button_cell.hyperlink = f"#{toc_sheet_name}!A1"
                
                # Set column width for button
                col_letter = button_cell.column_letter
                worksheet.column_dimensions[col_letter].width = 6
                
                button_count += 1
                self.logger.debug(f"Added TOC button to sheet '{sheet_name}'")
        
        self.logger.info(f"✅ Added {button_count} navigation buttons")
        return button_count

    # Analysis and Reporting Methods
    
    def create_pivot_table_from_data(self, source_sheet: str,
                                   target_sheet: str,
                                   rows: List[str],
                                   columns: List[str] = None,
                                   values: List[str] = None,
                                   aggfunc: str = 'sum') -> str:
        """
        Generate pivot table from source data and create in target sheet.
        
        Args:
            source_sheet (str): Source sheet name with raw data
            target_sheet (str): Target sheet name for pivot table
            rows (List[str]): Column names to use as row labels
            columns (List[str]): Column names to use as column labels  
            values (List[str]): Column names to aggregate
            aggfunc (str): Aggregation function ('sum', 'mean', 'count', 'max', 'min')
            
        Returns:
            str: Name of created pivot table sheet
            
        Raises:
            KeyError: If source sheet doesn't exist
            ValueError: If workbook not loaded or invalid parameters
        """
        if not self.workbook:
            raise ValueError("No workbook loaded - call load_workbook() first")
        
        if source_sheet not in self.workbook.sheetnames:
            raise KeyError(f"Source sheet '{source_sheet}' not found")
        
        self.logger.debug(f"Creating pivot table from '{source_sheet}' to '{target_sheet}'")
        
        # Extract data from source sheet
        source_ws = self.workbook[source_sheet]
        data = []
        headers = []
        
        # Get headers
        for cell in source_ws[1]:
            headers.append(cell.value if cell.value is not None else "")
        
        # Get data
        for row in source_ws.iter_rows(min_row=2, values_only=True):
            data.append(list(row))
        
        if not data:
            raise ValueError("No data found in source sheet")
        
        # Create DataFrame
        df = pd.DataFrame(data, columns=headers)
        
        # Validate pivot parameters
        for col in rows:
            if col not in headers:
                raise ValueError(f"Row field '{col}' not found in data")
        
        if columns:
            for col in columns:
                if col not in headers:
                    raise ValueError(f"Column field '{col}' not found in data")
        
        if values:
            for col in values:
                if col not in headers:
                    raise ValueError(f"Value field '{col}' not found in data")
        
        # Create pivot table using pandas
        pivot_df = pd.pivot_table(
            df, 
            index=rows,
            columns=columns,
            values=values,
            aggfunc=aggfunc,
            fill_value=0
        )
        
        # Create or clear target sheet
        if target_sheet in self.workbook.sheetnames:
            target_ws = self.workbook[target_sheet]
            target_ws.delete_rows(1, target_ws.max_row)
        else:
            target_ws = self.workbook.create_sheet(title=target_sheet)
        
        # Write pivot table to worksheet
        # Handle MultiIndex columns if present
        if isinstance(pivot_df.columns, pd.MultiIndex):
            # Write multi-level headers
            col_labels = pivot_df.columns.tolist()
            for col_idx, col_label in enumerate(col_labels, start=len(rows)+1):
                if isinstance(col_label, tuple):
                    header_text = " | ".join(str(x) for x in col_label if x != "")
                else:
                    header_text = str(col_label)
                target_ws.cell(row=1, column=col_idx, value=header_text)
        else:
            # Write single-level headers
            for col_idx, col_name in enumerate(pivot_df.columns, start=len(rows)+1):
                target_ws.cell(row=1, column=col_idx, value=str(col_name))
        
        # Write row index headers
        for row_level, row_name in enumerate(rows):
            target_ws.cell(row=1, column=row_level+1, value=row_name)
        
        # Write pivot data
        for row_idx, (index_values, row_data) in enumerate(pivot_df.iterrows(), start=2):
            # Write index values
            if isinstance(index_values, tuple):
                for col_idx, value in enumerate(index_values):
                    target_ws.cell(row=row_idx, column=col_idx+1, value=value)
            else:
                target_ws.cell(row=row_idx, column=1, value=index_values)
            
            # Write data values
            for col_idx, value in enumerate(row_data, start=len(rows)+1):
                target_ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Style the pivot table
        self._style_pivot_table(target_ws, len(rows), len(pivot_df.columns))
        
        self.logger.info(f"✅ Created pivot table in sheet '{target_sheet}' with {len(pivot_df)} rows")
        return target_sheet

    def _style_pivot_table(self, worksheet, index_cols: int, data_cols: int) -> None:
        """
        Apply styling to pivot table for better readability.
        
        Args:
            worksheet: Target worksheet to style
            index_cols (int): Number of index columns
            data_cols (int): Number of data columns
        """
        # Style headers
        header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for col in range(1, index_cols + data_cols + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
        
        # Style index columns
        index_font = Font(name='Calibri', size=10, bold=True)
        for row in range(2, worksheet.max_row + 1):
            for col in range(1, index_cols + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = index_font

    def filter_column_by_value_range(self, sheet_name: str,
                                   column_name: str,
                                   min_value: Optional[float] = None,
                                   max_value: Optional[float] = None,
                                   target_sheet: str = None) -> int:
        """
        Filter data based on value range in specified column.
        
        Args:
            sheet_name (str): Source sheet name
            column_name (str): Column to filter on
            min_value (Optional[float]): Minimum value (inclusive)
            max_value (Optional[float]): Maximum value (inclusive)
            target_sheet (str): Target sheet for filtered data, modifies source if None
            
        Returns:
            int: Number of rows in filtered result
            
        Raises:
            KeyError: If sheet or column doesn't exist
            ValueError: If workbook not loaded or invalid range
        """
        if not self.workbook:
            raise ValueError("No workbook loaded - call load_workbook() first")
        
        if sheet_name not in self.workbook.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found")
        
        if min_value is None and max_value is None:
            raise ValueError("At least one of min_value or max_value must be specified")
        
        self.logger.debug(f"Filtering column '{column_name}' in sheet '{sheet_name}'")
        
        # Extract data from source
        source_ws = self.workbook[sheet_name]
        data = []
        headers = []
        
        # Get headers
        for cell in source_ws[1]:
            headers.append(cell.value if cell.value is not None else "")
        
        if column_name not in headers:
            raise KeyError(f"Column '{column_name}' not found in sheet")
        
        col_idx = headers.index(column_name)
        
        # Get and filter data
        for row in source_ws.iter_rows(min_row=2, values_only=True):
            row_list = list(row)
            cell_value = row_list[col_idx]
            
            # Skip non-numeric values
            try:
                numeric_value = float(cell_value) if cell_value is not None else None
                if numeric_value is None:
                    continue
                    
                # Apply range filter
                passes_filter = True
                if min_value is not None and numeric_value < min_value:
                    passes_filter = False
                if max_value is not None and numeric_value > max_value:
                    passes_filter = False
                
                if passes_filter:
                    data.append(row_list)
                    
            except (ValueError, TypeError):
                continue  # Skip non-numeric values
        
        # Determine target worksheet
        if target_sheet:
            if target_sheet in self.workbook.sheetnames:
                target_ws = self.workbook[target_sheet]
                target_ws.delete_rows(1, target_ws.max_row)
            else:
                target_ws = self.workbook.create_sheet(title=target_sheet)
        else:
            target_ws = source_ws
            target_ws.delete_rows(2, target_ws.max_row)
        
        # Write filtered data
        if target_sheet or (not target_sheet):
            # Write headers if new sheet
            if target_sheet:
                for col_idx, header in enumerate(headers, 1):
                    target_ws.cell(row=1, column=col_idx, value=header)
            
            # Write filtered data
            start_row = 2 if target_sheet else 2
            for row_idx, row_data in enumerate(data, start_row):
                for col_idx, cell_value in enumerate(row_data, 1):
                    target_ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        result_count = len(data)
        range_desc = f"min={min_value}" if min_value is not None else ""
        if max_value is not None:
            range_desc += f", max={max_value}" if range_desc else f"max={max_value}"
        
        target_name = target_sheet if target_sheet else sheet_name
        self.logger.info(f"✅ Filtered {result_count} rows in '{target_name}' where '{column_name}' {range_desc}")
        
        return result_count

    def highlight_cells_by_value_range(self, sheet_name: str,
                                     column_name: str,
                                     min_value: Optional[float] = None,
                                     max_value: Optional[float] = None,
                                     highlight_color: str = "FFFF00") -> int:
        """
        Highlight cells in a column based on value range.
        
        Args:
            sheet_name (str): Sheet name to modify
            column_name (str): Column to highlight
            min_value (Optional[float]): Minimum value for highlighting
            max_value (Optional[float]): Maximum value for highlighting  
            highlight_color (str): Hex color code for highlighting
            
        Returns:
            int: Number of cells highlighted
            
        Raises:
            KeyError: If sheet or column doesn't exist
            ValueError: If workbook not loaded or invalid parameters
        """
        if not self.workbook:
            raise ValueError("No workbook loaded - call load_workbook() first")
        
        if sheet_name not in self.workbook.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found")
        
        if min_value is None and max_value is None:
            raise ValueError("At least one of min_value or max_value must be specified")
        
        worksheet = self.workbook[sheet_name]
        self.logger.debug(f"Highlighting cells in column '{column_name}' in sheet '{sheet_name}'")
        
        # Find column index
        headers = []
        for cell in worksheet[1]:
            headers.append(cell.value if cell.value is not None else "")
        
        if column_name not in headers:
            raise KeyError(f"Column '{column_name}' not found in sheet")
        
        col_idx = headers.index(column_name) + 1  # Excel columns are 1-indexed
        highlighted_count = 0
        
        # Create highlight fill
        highlight_fill = PatternFill(start_color=highlight_color, 
                                   end_color=highlight_color, 
                                   fill_type="solid")
        
        # Process each data row
        for row_num in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_num, column=col_idx)
            
            if cell.value is not None:
                try:
                    numeric_value = float(cell.value)
                    
                    # Check if value falls in highlight range
                    should_highlight = True
                    if min_value is not None and numeric_value < min_value:
                        should_highlight = False
                    if max_value is not None and numeric_value > max_value:
                        should_highlight = False
                    
                    if should_highlight:
                        cell.fill = highlight_fill
                        highlighted_count += 1
                        
                except (ValueError, TypeError):
                    continue  # Skip non-numeric values
        
        range_desc = f"min={min_value}" if min_value is not None else ""
        if max_value is not None:
            range_desc += f", max={max_value}" if range_desc else f"max={max_value}"
        
        self.logger.info(f"✅ Highlighted {highlighted_count} cells in column '{column_name}' where {range_desc}")
        return highlighted_count

    # Visualization and Dashboard Methods
    
    def create_chart_from_data(self, sheet_name: str,
                             chart_type: str,
                             data_range: str,
                             chart_title: str = "Chart",
                             x_axis_title: str = "X Axis",
                             y_axis_title: str = "Y Axis") -> str:
        """
        Create chart from data range in worksheet.
        
        Args:
            sheet_name (str): Sheet containing the data
            chart_type (str): Chart type ('bar', 'line', 'pie')
            data_range (str): Excel range for chart data (e.g., 'A1:C10')
            chart_title (str): Title for the chart
            x_axis_title (str): X-axis label
            y_axis_title (str): Y-axis label
            
        Returns:
            str: Chart identifier for reference
            
        Raises:
            KeyError: If sheet doesn't exist
            ValueError: If workbook not loaded or invalid chart type
        """
        if not self.workbook:
            raise ValueError("No workbook loaded - call load_workbook() first")
        
        if sheet_name not in self.workbook.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found")
        
        valid_chart_types = ['bar', 'line', 'pie']
        if chart_type not in valid_chart_types:
            raise ValueError(f"Invalid chart type '{chart_type}'. Must be one of: {valid_chart_types}")
        
        worksheet = self.workbook[sheet_name]
        self.logger.debug(f"Creating {chart_type} chart in sheet '{sheet_name}'")
        
        # Create chart based on type
        if chart_type == 'bar':
            chart = BarChart()
        elif chart_type == 'line':
            chart = LineChart()
        elif chart_type == 'pie':
            chart = PieChart()
        
        # Set chart properties
        chart.title = chart_title
        if chart_type != 'pie':  # Pie charts don't have axes
            chart.x_axis.title = x_axis_title
            chart.y_axis.title = y_axis_title
        
        # Add data to chart
        data_ref = Reference(worksheet, range_string=data_range)
        chart.add_data(data_ref, titles_from_data=True)
        
        # Position chart (try to find empty space)
        chart_anchor = self._find_chart_position(worksheet)
        worksheet.add_chart(chart, chart_anchor)
        
        chart_id = f"{sheet_name}_{chart_type}_{len(worksheet._charts)}"
        self.logger.info(f"✅ Created {chart_type} chart '{chart_title}' in sheet '{sheet_name}'")
        
        return chart_id

    def _find_chart_position(self, worksheet) -> str:
        """
        Find suitable position for chart placement.
        
        Args:
            worksheet: Target worksheet
            
        Returns:
            str: Excel cell reference for chart anchor
        """
        # Try to place chart to the right of data
        max_col = worksheet.max_column
        chart_col = max_col + 2 if max_col < 20 else 1  # Place at column A if too far right
        
        # Find row with enough space
        chart_row = 1
        
        return f"{chr(ord('A') + chart_col - 1)}{chart_row}"

    def apply_conditional_formatting(self, sheet_name: str,
                                   data_range: str,
                                   rule_type: str,
                                   format_style: str = "3_color_scale") -> int:
        """
        Apply conditional formatting to data range.
        
        Args:
            sheet_name (str): Sheet to format
            data_range (str): Excel range to format (e.g., 'B2:E10')
            rule_type (str): Type of formatting ('color_scale', 'above_average', 'below_average')
            format_style (str): Style of formatting
            
        Returns:
            int: Number of cells formatted
            
        Raises:
            KeyError: If sheet doesn't exist
            ValueError: If workbook not loaded or invalid parameters
        """
        if not self.workbook:
            raise ValueError("No workbook loaded - call load_workbook() first")
        
        if sheet_name not in self.workbook.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found")
        
        worksheet = self.workbook[sheet_name]
        self.logger.debug(f"Applying conditional formatting to range '{data_range}' in sheet '{sheet_name}'")
        
        if rule_type == "color_scale":
            # Apply 3-color scale formatting
            rule = ColorScaleRule(
                start_type='min', start_color='FF0000',  # Red for low values
                mid_type='percentile', mid_value=50, mid_color='FFFF00',  # Yellow for mid
                end_type='max', end_color='00FF00'  # Green for high values
            )
            worksheet.conditional_formatting.add(data_range, rule)
            
        elif rule_type == "above_average":
            rule = CellIsRule(
                operator='aboveAverage',
                fill=PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            )
            worksheet.conditional_formatting.add(data_range, rule)
            
        elif rule_type == "below_average":
            rule = CellIsRule(
                operator='belowAverage',
                fill=PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
            )
            worksheet.conditional_formatting.add(data_range, rule)
        
        else:
            raise ValueError(f"Unknown rule_type '{rule_type}'")
        
        # Count cells in range for reporting
        from openpyxl.utils import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(data_range)
        cell_count = (max_col - min_col + 1) * (max_row - min_row + 1)
        
        self.logger.info(f"✅ Applied {rule_type} conditional formatting to {cell_count} cells")
        return cell_count

    def create_summary_sheet(self, source_sheets: List[str],
                           summary_sheet_name: str = "Summary",
                           include_charts: bool = True) -> str:
        """
        Create summary sheet with KPIs and charts from multiple source sheets.
        
        Args:
            source_sheets (List[str]): List of sheets to summarize
            summary_sheet_name (str): Name for summary sheet
            include_charts (bool): Whether to include summary charts
            
        Returns:
            str: Name of created summary sheet
            
        Raises:
            KeyError: If any source sheet doesn't exist
            ValueError: If workbook not loaded
        """
        if not self.workbook:
            raise ValueError("No workbook loaded - call load_workbook() first")
        
        # Validate source sheets exist
        missing_sheets = [name for name in source_sheets if name not in self.workbook.sheetnames]
        if missing_sheets:
            raise KeyError(f"Source sheets not found: {missing_sheets}")
        
        self.logger.debug(f"Creating summary sheet from {len(source_sheets)} source sheets")
        
        # Create or clear summary sheet
        if summary_sheet_name in self.workbook.sheetnames:
            summary_ws = self.workbook[summary_sheet_name]
            summary_ws.delete_rows(1, summary_ws.max_row)
        else:
            summary_ws = self.workbook.create_sheet(title=summary_sheet_name)
        
        # Create summary title
        title_font = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        summary_ws.cell(row=1, column=1, value="DATA SUMMARY")
        summary_ws.cell(row=1, column=1).font = title_font
        summary_ws.cell(row=1, column=1).fill = title_fill
        summary_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        
        # Generate KPIs section
        current_row = 3
        kpi_font = Font(name='Calibri', size=12, bold=True)
        
        summary_ws.cell(row=current_row, column=1, value="KEY PERFORMANCE INDICATORS")
        summary_ws.cell(row=current_row, column=1).font = kpi_font
        current_row += 2
        
        # Calculate KPIs for each source sheet
        for sheet_name in source_sheets:
            worksheet = self.workbook[sheet_name]
            
            # Calculate basic statistics
            row_count = worksheet.max_row - 1 if worksheet.max_row > 1 else 0  # Exclude header
            col_count = worksheet.max_column
            
            # Sheet summary
            summary_ws.cell(row=current_row, column=1, value=f"Sheet: {sheet_name}")
            summary_ws.cell(row=current_row, column=2, value=f"Rows: {row_count}")
            summary_ws.cell(row=current_row, column=3, value=f"Columns: {col_count}")
            
            # Calculate numeric summaries for numeric columns
            numeric_summaries = self._calculate_numeric_summaries(worksheet)
            col_offset = 4
            for metric_name, value in numeric_summaries.items():
                summary_ws.cell(row=current_row, column=col_offset, value=f"{metric_name}: {value}")
                col_offset += 1
            
            current_row += 1
        
        # Add charts if requested
        if include_charts and len(source_sheets) > 1:
            current_row += 2
            summary_ws.cell(row=current_row, column=1, value="VISUAL SUMMARY")
            summary_ws.cell(row=current_row, column=1).font = kpi_font
            
            # Create comparison chart
            self._create_comparison_chart(summary_ws, source_sheets, current_row + 2)
        
        # Auto-adjust column widths
        for column in summary_ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            summary_ws.column_dimensions[column_letter].width = adjusted_width
        
        self.logger.info(f"✅ Created summary sheet '{summary_sheet_name}' for {len(source_sheets)} sheets")
        return summary_sheet_name

    def _calculate_numeric_summaries(self, worksheet) -> Dict[str, float]:
        """
        Calculate numeric summaries for worksheet data.
        
        Args:
            worksheet: Source worksheet
            
        Returns:
            Dict[str, float]: Dictionary of metric names to values
        """
        summaries = {}
        
        # Find numeric columns and calculate basic stats
        numeric_values = []
        
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            for cell_value in row:
                if cell_value is not None:
                    try:
                        numeric_values.append(float(cell_value))
                    except (ValueError, TypeError):
                        continue
        
        if numeric_values:
            summaries['Total'] = sum(numeric_values)
            summaries['Count'] = len(numeric_values)
            summaries['Average'] = sum(numeric_values) / len(numeric_values)
            summaries['Max'] = max(numeric_values)
            summaries['Min'] = min(numeric_values)
        
        return summaries

    def _create_comparison_chart(self, worksheet, source_sheets: List[str], start_row: int) -> None:
        """
        Create comparison chart in summary worksheet.
        
        Args:
            worksheet: Target worksheet
            source_sheets (List[str]): Source sheets to compare
            start_row (int): Starting row for chart data
        """
        # Create simple row count comparison chart
        chart = BarChart()
        chart.title = "Sheet Comparison - Row Counts"
        chart.x_axis.title = "Sheets"
        chart.y_axis.title = "Number of Rows"
        
        # Prepare chart data
        for idx, sheet_name in enumerate(source_sheets):
            source_ws = self.workbook[sheet_name]
            row_count = source_ws.max_row - 1 if source_ws.max_row > 1 else 0
            
            # Write chart data
            worksheet.cell(row=start_row + idx, column=1, value=sheet_name)
            worksheet.cell(row=start_row + idx, column=2, value=row_count)
        
        # Add data to chart
        data_range = f"B{start_row}:B{start_row + len(source_sheets) - 1}"
        categories_range = f"A{start_row}:A{start_row + len(source_sheets) - 1}"
        
        data_ref = Reference(worksheet, range_string=data_range)
        cats_ref = Reference(worksheet, range_string=categories_range)
        
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(cats_ref)
        
        # Position and add chart
        chart_anchor = f"D{start_row}"
        worksheet.add_chart(chart, chart_anchor)

    def format_chart_elements(self, sheet_name: str,
                            chart_index: int = 0,
                            title_font_size: int = 14,
                            axis_font_size: int = 10,
                            show_gridlines: bool = True) -> bool:
        """
        Format chart elements for better appearance.
        
        Args:
            sheet_name (str): Sheet containing the chart
            chart_index (int): Index of chart to format (0 for first chart)
            title_font_size (int): Font size for chart title
            axis_font_size (int): Font size for axis labels
            show_gridlines (bool): Whether to show gridlines
            
        Returns:
            bool: True if chart was formatted successfully
            
        Raises:
            KeyError: If sheet doesn't exist
            ValueError: If workbook not loaded
        """
        if not self.workbook:
            raise ValueError("No workbook loaded - call load_workbook() first")
        
        if sheet_name not in self.workbook.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found")
        
        worksheet = self.workbook[sheet_name]
        
        try:
            # Access chart by index
            if len(worksheet._charts) <= chart_index:
                self.logger.warning(f"Chart index {chart_index} not found in sheet '{sheet_name}'")
                return False
            
            chart = worksheet._charts[chart_index]
            
            # Format title
            if hasattr(chart, 'title') and chart.title:
                chart.title.tx.rich.p[0].r[0].rPr.sz = title_font_size * 100  # OpenPyXL uses EMUs
            
            # Format axes if available
            if hasattr(chart, 'x_axis') and chart.x_axis:
                chart.x_axis.txPr = chart.x_axis.txPr or {}
                
            if hasattr(chart, 'y_axis') and chart.y_axis:
                chart.y_axis.txPr = chart.y_axis.txPr or {}
            
            # Gridlines
            if hasattr(chart, 'y_axis') and chart.y_axis:
                chart.y_axis.majorGridlines = None if not show_gridlines else chart.y_axis.majorGridlines
            
            self.logger.info(f"✅ Formatted chart {chart_index} in sheet '{sheet_name}'")
            return True
            
        except Exception as e:
            self.logger.error(f"Failed to format chart: {e}")
            return False


class ExcelConfigV1:
    """
    Configuration class for Excel automation operations.
    
    Provides structured configuration management for Excel processing
    tasks with JSON serialization support and validation.
    """
    
    def __init__(self) -> None:
        """Initialize Excel configuration with default values."""
        self.data_cleaning: Dict[str, Any] = {
            "remove_duplicates": True,
            "standardize_dates": True,
            "clean_text_casing": "title",
            "normalize_columns": True
        }
        
        self.analysis: Dict[str, Any] = {
            "create_pivot_tables": True,
            "generate_summaries": True,
            "apply_filters": True,
            "highlight_outliers": True
        }
        
        self.visualization: Dict[str, Any] = {
            "create_charts": True,
            "apply_conditional_formatting": True,
            "create_dashboards": False,
            "chart_types": ["bar", "line", "pie"]
        }
        
        self.navigation: Dict[str, Any] = {
            "create_toc": True,
            "add_nav_buttons": True,
            "toc_threshold": 6  # Create TOC if more than 6 sheets
        }
        
        self.logger = get_logger(f"{__name__}.{self.__class__.__name__}")
    
    def load_from_json(self, config_path: str) -> 'ExcelConfigV1':
        """
        Load configuration from JSON file.
        
        Args:
            config_path (str): Path to JSON configuration file
            
        Returns:
            ExcelConfigV1: Self for method chaining
            
        Raises:
            FileNotFoundError: If config file doesn't exist
            ValueError: If JSON is invalid
        """
        try:
            with open(config_path, 'r') as f:
                import json
                config_data = json.load(f)
            
            # Update configuration sections
            if 'data_cleaning' in config_data:
                self.data_cleaning.update(config_data['data_cleaning'])
            
            if 'analysis' in config_data:
                self.analysis.update(config_data['analysis'])
            
            if 'visualization' in config_data:
                self.visualization.update(config_data['visualization'])
            
            if 'navigation' in config_data:
                self.navigation.update(config_data['navigation'])
            
            self.logger.info(f"✅ Loaded configuration from: {config_path}")
            return self
            
        except FileNotFoundError:
            self.logger.error(f"Configuration file not found: {config_path}")
            raise
        except json.JSONDecodeError as e:
            self.logger.error(f"Invalid JSON in configuration file: {e}")
            raise ValueError(f"Invalid JSON configuration: {e}")
    
    def save_to_json(self, config_path: str) -> str:
        """
        Save current configuration to JSON file.
        
        Args:
            config_path (str): Path to save configuration file
            
        Returns:
            str: Path where configuration was saved
        """
        import json
        
        config_data = {
            "data_cleaning": self.data_cleaning,
            "analysis": self.analysis,
            "visualization": self.visualization,
            "navigation": self.navigation
        }
        
        try:
            # Ensure directory exists
            Path(config_path).parent.mkdir(parents=True, exist_ok=True)
            
            with open(config_path, 'w') as f:
                json.dump(config_data, f, indent=2)
            
            self.logger.info(f"✅ Configuration saved to: {config_path}")
            return config_path
            
        except Exception as e:
            self.logger.error(f"Failed to save configuration: {e}")
            raise


def create_example_config() -> str:
    """
    Create an example configuration file for Excel automation.
    
    Returns:
        str: Path to created example configuration file
    """
    config = ExcelConfigV1()
    
    # Customize example settings
    config.data_cleaning.update({
        "date_format": "MM/DD/YYYY",
        "currency_format": "$#,##0.00",
        "text_columns": ["Name", "Description", "Category"],
        "date_columns": ["Date", "Created", "Modified"]
    })
    
    config.analysis.update({
        "pivot_settings": {
            "default_aggfunc": "sum",
            "include_totals": True
        },
        "filter_settings": {
            "highlight_color": "FFFF00",
            "conditional_rules": ["above_average", "color_scale"]
        }
    })
    
    config.visualization.update({
        "chart_settings": {
            "default_width": 15,
            "default_height": 10,
            "title_font_size": 14,
            "show_gridlines": True
        }
    })
    
    # Save example config
    example_path = "excel_automation_config_example.json"
    return config.save_to_json(example_path)


# Utility functions for common Excel operations

def auto_clean_excel_file(filepath: str, 
                         config: Optional[ExcelConfigV1] = None,
                         output_path: Optional[str] = None) -> str:
    """
    Automatically clean an Excel file using default or provided configuration.
    
    Args:
        filepath (str): Path to Excel file to clean
        config (Optional[ExcelConfigV1]): Configuration object, uses defaults if None
        output_path (Optional[str]): Output path, modifies original if None
        
    Returns:
        str: Path to cleaned Excel file
        
    Raises:
        FileNotFoundError: If input file doesn't exist
        ValueError: If file cannot be processed
    """
    if not Path(filepath).exists():
        raise FileNotFoundError(f"Excel file not found: {filepath}")
    
    # Use default config if none provided
    if config is None:
        config = ExcelConfigV1()
    
    logger.info(f"Auto-cleaning Excel file: {filepath}")
    
    # Initialize processor
    processor = ExcelProcessorV1(filepath, debug=os.getenv("DEBUG") == "1")
    processor.load_workbook()
    
    # Apply data cleaning operations
    if config.data_cleaning.get("remove_duplicates", True):
        for sheet_name in processor.workbook.sheetnames:
            try:
                removed = processor.remove_duplicates_from_sheet(sheet_name)
                if removed > 0:
                    logger.debug(f"Removed {removed} duplicates from '{sheet_name}'")
            except Exception as e:
                logger.warning(f"Could not clean duplicates from '{sheet_name}': {e}")
    
    if config.data_cleaning.get("normalize_columns", True):
        for sheet_name in processor.workbook.sheetnames:
            try:
                mappings = processor.normalize_column_names(sheet_name)
                if mappings:
                    logger.debug(f"Normalized {len(mappings)} columns in '{sheet_name}'")
            except Exception as e:
                logger.warning(f"Could not normalize columns in '{sheet_name}': {e}")
    
    # Apply navigation enhancements if multiple sheets
    sheet_count = len(processor.workbook.sheetnames)
    if (sheet_count > config.navigation.get("toc_threshold", 6) and 
        config.navigation.get("create_toc", True)):
        
        try:
            processor.create_table_of_contents()
            if config.navigation.get("add_nav_buttons", True):
                processor.add_navigation_buttons()
            logger.debug("Added navigation features for multi-sheet workbook")
        except Exception as e:
            logger.warning(f"Could not add navigation features: {e}")
    
    # Save processed file
    output_file = processor.save_workbook(output_path)
    logger.info(f"✅ Auto-cleaning completed: {output_file}")
    
    return output_file


def batch_process_excel_files(directory_path: str,
                            config: Optional[ExcelConfigV1] = None,
                            output_directory: Optional[str] = None) -> List[str]:
    """
    Process multiple Excel files in a directory.
    
    Args:
        directory_path (str): Directory containing Excel files
        config (Optional[ExcelConfigV1]): Configuration for processing
        output_directory (Optional[str]): Output directory, modifies originals if None
        
    Returns:
        List[str]: List of processed file paths
        
    Raises:
        FileNotFoundError: If directory doesn't exist
    """
    dir_path = Path(directory_path)
    if not dir_path.exists():
        raise FileNotFoundError(f"Directory not found: {directory_path}")
    
    # Find Excel files
    excel_files = []
    for pattern in ['*.xlsx', '*.xls', '*.xlsm']:
        excel_files.extend(dir_path.glob(pattern))
    
    if not excel_files:
        logger.warning(f"No Excel files found in directory: {directory_path}")
        return []
    
    logger.info(f"Found {len(excel_files)} Excel files to process")
    
    processed_files = []
    
    for excel_file in excel_files:
        try:
            # Determine output path
            if output_directory:
                output_path = Path(output_directory) / f"processed_{excel_file.name}"
            else:
                output_path = None
            
            # Process file
            result_path = auto_clean_excel_file(str(excel_file), config, str(output_path) if output_path else None)
            processed_files.append(result_path)
            
            logger.debug(f"Processed: {excel_file.name}")
            
        except Exception as e:
            logger.error(f"Failed to process {excel_file.name}: {e}")
            continue
    
    logger.info(f"✅ Batch processing completed: {len(processed_files)}/{len(excel_files)} files processed")
    return processed_files