# tests/test_excel_processor.py

"""
Comprehensive test suite for Excel processing functionality.

This module provides thorough testing of the ExcelProcessorV1 class and
related utilities, ensuring all features work correctly and handle edge cases.

Tests cover:
- Data cleaning and transformation operations
- Analysis and reporting features
- Visualization and dashboard creation
- Configuration management
- Error handling and edge cases
- CLI integration

Follows Framework0 testing patterns with pytest, fixtures, and mocks.
"""

import pytest
import pandas as pd
import openpyxl
from pathlib import Path
import tempfile
import json
import os
from unittest.mock import Mock, patch

# Import modules under test
import sys
from typing import Any, Dict, List, Optional, Union
from pathlib import Path

# Add project root to path for imports
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from analysis.excel_processor import ExcelProcessorV1, ExcelConfigV1, auto_clean_excel_file

# Import CLI directly to avoid circular import issues
try:
    sys.path.insert(0, str(project_root / 'cli'))
    from excel_automation import ExcelAutomationCLI
except ImportError:
    # Skip CLI tests if import fails
    ExcelAutomationCLI = None


class TestExcelProcessorV1:
    """Test suite for ExcelProcessorV1 class."""
    
    @pytest.fixture
def sample_excel_file(self) -> Any:
    """Create temporary Excel file with test data."""
    # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_file.close()
        
        # Create test data
        sales_data = pd.DataFrame({
            'Date': ['2024-01-15', '2024-01-16', '2024-01-15', '2024-01-17'],  # One duplicate
            'Region': ['North', 'South', 'North', 'East'],
            'product_name': ['Widget A', 'Widget B', 'Widget A', 'Widget C'],  # Needs normalization
            'Sales': [1500, 2000, 1500, 1200],
            'cost': [800, 1100, 800, 600]  # Needs normalization
        })
        
        customers_data = pd.DataFrame({
            'Customer_ID': [1, 2, 3, 4],
            'customer_name': ['ABC Corp', 'xyz inc', 'DEF LLC', 'ghi co'],  # Mixed case
            'Region': ['North', 'South', 'North', 'East'],
            'Total_Purchase': [5000, 7500, 3200, 8900]
        })
        
        # Write test data to Excel file
        with pd.ExcelWriter(temp_file.name, engine='openpyxl') as writer:
            sales_data.to_excel(writer, sheet_name='Sales_Data', index=False)
            customers_data.to_excel(writer, sheet_name='Customers', index=False)
        
        yield temp_file.name
        
        # Cleanup
        Path(temp_file.name).unlink(missing_ok=True)
    
def test_processor_initialization(self, sample_excel_file -> Any: Any):
        """Test processor initialization with valid file."""
        # Initialize processor
        processor = ExcelProcessorV1(sample_excel_file, debug=True)
        
        # Verify initialization
        assert processor.filepath == sample_excel_file
        assert processor.debug is True
        assert processor.workbook is None
        assert processor.original_sheets == []
        assert processor.logger is not None
    
def test_load_workbook_existing_file(self, sample_excel_file -> Any: Any):
        """Test loading existing Excel workbook."""
        # Initialize and load
        processor = ExcelProcessorV1(sample_excel_file)
        workbook = processor.load_workbook()
        
        # Verify workbook loaded
        assert workbook is not None
        assert isinstance(workbook, openpyxl.Workbook)
        assert len(processor.original_sheets) == 2
        assert 'Sales_Data' in processor.original_sheets
        assert 'Customers' in processor.original_sheets
    
def test_load_workbook_new_file(self) -> Any:
    """Test creating new workbook for non-existent file."""
    # Use non-existent file path
        temp_path = tempfile.mktemp(suffix='.xlsx')
        
        try:
            # Initialize processor
            processor = ExcelProcessorV1(temp_path)
            workbook = processor.load_workbook()
            
            # Verify new workbook created
            assert workbook is not None
            assert isinstance(workbook, openpyxl.Workbook)
            assert len(processor.original_sheets) == 0
            
        finally:
            # Cleanup if file was created
            Path(temp_path).unlink(missing_ok=True)
    
def test_remove_duplicates(self, sample_excel_file -> Any: Any):
        """Test duplicate removal functionality."""
        # Initialize and load workbook
        processor = ExcelProcessorV1(sample_excel_file)
        processor.load_workbook()
        
        # Remove duplicates from Sales_Data sheet
        removed_count = processor.remove_duplicates_from_sheet('Sales_Data')
        
        # Verify duplicates removed
        assert removed_count == 1  # Should remove 1 duplicate row
        
        # Verify sheet content
        worksheet = processor.workbook['Sales_Data']
        data_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == 3  # Original 4 rows minus 1 duplicate
    
def test_normalize_column_names(self, sample_excel_file -> Any: Any):
        """Test column name normalization."""
        # Initialize and load workbook
        processor = ExcelProcessorV1(sample_excel_file)
        processor.load_workbook()
        
        # Normalize columns in Sales_Data sheet
        mappings = processor.normalize_column_names('Sales_Data')
        
        # Verify normalizations occurred
        assert 'product_name' in mappings
        assert mappings['product_name'] == 'Product_Name'
        assert 'cost' in mappings
        assert mappings['cost'] == 'Cost'
        
        # Verify changes in worksheet
        worksheet = processor.workbook['Sales_Data']
        headers = [cell.value for cell in worksheet[1]]
        assert 'Product_Name' in headers
        assert 'Cost' in headers
        assert 'product_name' not in headers
        assert 'cost' not in headers
    
def test_clean_text_casing(self, sample_excel_file -> Any: Any):
        """Test text casing standardization."""
        # Initialize and load workbook
        processor = ExcelProcessorV1(sample_excel_file)
        processor.load_workbook()
        
        # Clean text casing in customer names
        processed_count = processor.clean_text_casing(
            'Customers', 
            ['customer_name'], 
            'title'
        )
        
        # Verify text processing occurred
        assert processed_count > 0
        
        # Check specific values were transformed
        worksheet = processor.workbook['Customers']
        customer_names = [row[1] for row in worksheet.iter_rows(min_row=2, max_col=2, values_only=True)]
        
        # Should be title cased
        expected_names = ['ABC Corp', 'Xyz Inc', 'DEF LLC', 'Ghi Co']
        assert customer_names == expected_names
    
def test_create_table_of_contents(self, sample_excel_file -> Any: Any):
        """Test table of contents creation."""
        # Initialize and load workbook
        processor = ExcelProcessorV1(sample_excel_file)
        processor.load_workbook()
        
        # Create table of contents
        toc_sheet_name = processor.create_table_of_contents()
        
        # Verify TOC created
        assert toc_sheet_name == "Table of Contents"
        assert toc_sheet_name in processor.workbook.sheetnames
        
        # Verify TOC content
        toc_sheet = processor.workbook[toc_sheet_name]
        assert toc_sheet.cell(1, 1).value == "TABLE OF CONTENTS"
        
        # Should have links to other sheets
        sheet_links = []
        for row in toc_sheet.iter_rows(min_row=4, max_col=1, values_only=True):
            if row[0] and row[0] != toc_sheet_name:
                sheet_links.append(row[0])
        
        assert 'Sales_Data' in sheet_links
        assert 'Customers' in sheet_links
    
def test_add_navigation_buttons(self, sample_excel_file -> Any: Any):
        """Test navigation button addition."""
        # Initialize and load workbook
        processor = ExcelProcessorV1(sample_excel_file)
        processor.load_workbook()
        
        # Create TOC first
        toc_name = processor.create_table_of_contents()
        
        # Add navigation buttons
        button_count = processor.add_navigation_buttons(toc_name)
        
        # Verify buttons added
        assert button_count == 2  # One for each non-TOC sheet
        
        # Check buttons exist in sheets
        for sheet_name in ['Sales_Data', 'Customers']:
            worksheet = processor.workbook[sheet_name]
            # Find TOC button (should be in top-right area)
            toc_button_found = False
            for row in worksheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    if cell.value == "TOC":
                        toc_button_found = True
                        break
            assert toc_button_found, f"TOC button not found in sheet {sheet_name}"
    
def test_invalid_sheet_operations(self, sample_excel_file -> Any: Any):
        """Test error handling for invalid sheet operations."""
        # Initialize and load workbook
        processor = ExcelProcessorV1(sample_excel_file)
        processor.load_workbook()
        
        # Test operations on non-existent sheet
        with pytest.raises(KeyError):
            processor.remove_duplicates_from_sheet('NonExistentSheet')
        
        with pytest.raises(KeyError):
            processor.normalize_column_names('NonExistentSheet')
        
        with pytest.raises(KeyError):
            processor.clean_text_casing('NonExistentSheet', ['Column'], 'title')
    
def test_save_workbook(self, sample_excel_file -> Any: Any):
        """Test workbook saving functionality."""
        # Initialize and load workbook
        processor = ExcelProcessorV1(sample_excel_file)
        processor.load_workbook()
        
        # Make some changes
        processor.normalize_column_names('Sales_Data')
        
        # Save to new location
        output_path = tempfile.mktemp(suffix='_output.xlsx')
        try:
            saved_path = processor.save_workbook(output_path)
            
            # Verify file saved
            assert saved_path == output_path
            assert Path(output_path).exists()
            
            # Verify content preserved
            saved_workbook = openpyxl.load_workbook(output_path)
            assert len(saved_workbook.sheetnames) >= 2
            
        finally:
            # Cleanup
            Path(output_path).unlink(missing_ok=True)


class TestExcelConfigV1:
    """Test suite for ExcelConfigV1 class."""
    
def test_config_initialization(self) -> Any:
    """Test configuration initialization with defaults."""
    # Create config
        config = ExcelConfigV1()
        
        # Verify default settings
        assert config.data_cleaning['remove_duplicates'] is True
        assert config.data_cleaning['standardize_dates'] is True
        assert config.data_cleaning['clean_text_casing'] == 'title'
        assert config.data_cleaning['normalize_columns'] is True
        
        assert config.analysis['create_pivot_tables'] is True
        assert config.analysis['generate_summaries'] is True
        
        assert config.visualization['create_charts'] is True
        assert 'bar' in config.visualization['chart_types']
        
        assert config.navigation['create_toc'] is True
        assert config.navigation['toc_threshold'] == 6
    
def test_config_json_serialization(self) -> Any:
    """Test configuration JSON save/load functionality."""
    # Create config with custom settings
        config = ExcelConfigV1()
        config.data_cleaning['remove_duplicates'] = False
        config.navigation['toc_threshold'] = 3
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False)
        temp_file.close()
        
        try:
            # Save configuration
            saved_path = config.save_to_json(temp_file.name)
            assert saved_path == temp_file.name
            assert Path(temp_file.name).exists()
            
            # Load configuration
            new_config = ExcelConfigV1()
            new_config.load_from_json(temp_file.name)
            
            # Verify settings preserved
            assert new_config.data_cleaning['remove_duplicates'] is False
            assert new_config.navigation['toc_threshold'] == 3
            
        finally:
            # Cleanup
            Path(temp_file.name).unlink(missing_ok=True)
    
def test_config_invalid_json(self) -> Any:
    """Test handling of invalid JSON configuration."""
    # Create invalid JSON file
        temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False)
        temp_file.write('{ invalid json content }')
        temp_file.close()
        
        try:
            # Attempt to load invalid JSON
            config = ExcelConfigV1()
            with pytest.raises(ValueError):
                config.load_from_json(temp_file.name)
                
        finally:
            # Cleanup
            Path(temp_file.name).unlink(missing_ok=True)


@pytest.mark.skipif(ExcelAutomationCLI is None, reason="CLI module not available")
class TestExcelAutomationCLI:
    """Test suite for Excel Automation CLI."""
    
    @pytest.fixture
def cli(self) -> Any:
    """Create CLI instance for testing."""
    return ExcelAutomationCLI()
    
    @pytest.fixture
def sample_excel_file(self) -> Any:
    """Create temporary Excel file for CLI testing."""
    # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_file.close()
        
        # Create simple test data
        test_data = pd.DataFrame({
            'Name': ['Product A', 'Product B', 'Product C'],
            'Value': [100, 200, 150]
        })
        
        with pd.ExcelWriter(temp_file.name, engine='openpyxl') as writer:
            test_data.to_excel(writer, sheet_name='Data', index=False)
        
        yield temp_file.name
        
        # Cleanup
        Path(temp_file.name).unlink(missing_ok=True)
    
def test_cli_initialization(self, cli -> Any: Any):
        """Test CLI initialization."""
        # Verify CLI components initialized
        assert cli.parser is not None
        assert cli.config is None
        assert cli.debug is False
        assert cli.logger is not None
    
def test_cli_help_output(self, cli -> Any: Any):
        """Test CLI help output."""
        # Test main help
        with pytest.raises(SystemExit) as exc_info:
            cli.run(['--help'])
        assert exc_info.value.code == 0
    
def test_cli_create_config_command(self, cli -> Any: Any):
        """Test create-config CLI command."""
        # Create temporary output file
        output_file = tempfile.mktemp(suffix='.json')
        
        try:
            # Run create-config command
            result = cli.run(['create-config', '--template', 'basic', '--output', output_file])
            
            # Verify success
            assert result == 0
            assert Path(output_file).exists()
            
            # Verify config content
            with open(output_file, 'r') as f:
                config_data = json.load(f)
            
            assert 'data_cleaning' in config_data
            assert 'analysis' in config_data
            assert 'visualization' in config_data
            assert 'navigation' in config_data
            
        finally:
            # Cleanup
            Path(output_file).unlink(missing_ok=True)
    
def test_cli_auto_process_command(self, cli -> Any: Any, sample_excel_file: Any):
        """Test auto-process CLI command."""
        # Create temporary output file
        output_file = tempfile.mktemp(suffix='_processed.xlsx')
        
        try:
            # Run auto-process command
            result = cli.run(['--output', output_file, 'auto-process', sample_excel_file])
            
            # Verify success
            assert result == 0
            assert Path(output_file).exists()
            
            # Verify processed file has expected content
            processed_workbook = openpyxl.load_workbook(output_file)
            assert 'Data' in processed_workbook.sheetnames
            
        finally:
            # Cleanup
            Path(output_file).unlink(missing_ok=True)
    
def test_cli_nonexistent_file_error(self, cli -> Any: Any):
        """Test CLI error handling for non-existent files."""
        # Run command with non-existent file
        result = cli.run(['auto-process', '/nonexistent/file.xlsx'])
        
        # Should return error code
        assert result == 1


class TestUtilityFunctions:
    """Test suite for utility functions."""
    
    @pytest.fixture
def sample_excel_file(self) -> Any:
    """Create temporary Excel file for utility testing."""
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_file.close()
        
        # Create test data with duplicates and mixed formatting
        test_data = pd.DataFrame({
            'item_name': ['product a', 'PRODUCT B', 'Product C', 'product a'],  # Mixed case, duplicate
            'category': ['electronics', 'CLOTHING', 'books', 'electronics'],
            'price': [29.99, 149.50, 15.99, 29.99],
            'date_added': ['2024-01-15', '2024/01/16', '15-Jan-2024', '2024-01-15']  # Mixed date formats
        })
        
        with pd.ExcelWriter(temp_file.name, engine='openpyxl') as writer:
            test_data.to_excel(writer, sheet_name='Products', index=False)
        
        yield temp_file.name
        
        Path(temp_file.name).unlink(missing_ok=True)
    
def test_auto_clean_excel_file(self, sample_excel_file -> Any: Any):
        """Test auto_clean_excel_file utility function."""
        # Create output path
        output_path = tempfile.mktemp(suffix='_cleaned.xlsx')
        
        try:
            # Run auto-cleaning
            result_path = auto_clean_excel_file(sample_excel_file, output_path=output_path)
            
            # Verify cleaning completed
            assert result_path == output_path
            assert Path(output_path).exists()
            
            # Verify cleaning effects
            cleaned_workbook = openpyxl.load_workbook(output_path)
            worksheet = cleaned_workbook['Products']
            
            # Should have fewer rows due to duplicate removal
            data_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
            assert len(data_rows) == 3  # 4 original minus 1 duplicate
            
            # Column names should be normalized
            headers = [cell.value for cell in worksheet[1]]
            assert 'Item_Name' in headers  # 'item_name' normalized
            
        finally:
            Path(output_path).unlink(missing_ok=True)
    
def test_auto_clean_with_custom_config(self, sample_excel_file -> Any: Any):
        """Test auto-cleaning with custom configuration."""
        # Create custom configuration
        config = ExcelConfigV1()
        config.data_cleaning['remove_duplicates'] = False  # Disable duplicate removal
        
        # Create output path
        output_path = tempfile.mktemp(suffix='_custom_cleaned.xlsx')
        
        try:
            # Run auto-cleaning with custom config
            result_path = auto_clean_excel_file(sample_excel_file, config=config, output_path=output_path)
            
            # Verify processing completed
            assert result_path == output_path
            assert Path(output_path).exists()
            
            # Verify duplicates were NOT removed (due to config)
            cleaned_workbook = openpyxl.load_workbook(output_path)
            worksheet = cleaned_workbook['Products']
            data_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
            assert len(data_rows) == 4  # All original rows preserved
            
        finally:
            Path(output_path).unlink(missing_ok=True)


if __name__ == '__main__':
    # Run tests if executed directly
    pytest.main([__file__, '-v'])